from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import time
import datetime
from openpyxl import Workbook

class AirportButtonError(Exception):
    pass

class DateError(Exception):
    pass

# Your departure and destination details
departure_airport = "RIX"
destination_airport = "AMS"
departure_date = "2024-3-11"
return_date = "2024-3-13"

# Set up the Chrome webdriver
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(options=options)

# Navigate to the AirBaltic website
url = "https://www.airbaltic.com/en-LV/index"
driver.get(url)

def accept_cookies(driver):
    try:
        cookie_popup = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "cookie-alert"))
        )
        accept_button = cookie_popup.find_element(By.XPATH, "//button[contains(text(), 'Accept')]")
        accept_button.click()
    except TimeoutException:
        print("No cookie consent popup found or it was not displayed within the timeout.")

def set_departure_airport(driver, airport_code):
    try:
        airport_input = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'fsf-typeahead') and contains(@class, 'is-index')][1]//input"))
        )
        airport_input.clear()
        airport_input.send_keys(airport_code)
        airport_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "airport"))
        )
        airport_button.click()
    except TimeoutException:
        raise AirportButtonError(f"Departure option '{airport_code}' button not found or not clickable within the timeout.")


def set_destination_airport(driver, airport_code):
    destination_input = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'fsf-typeahead') and contains(@class, 'is-index')][2]//input"))
    )
    destination_input.clear()
    destination_input.send_keys(destination_airport)
    try:
        destination_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "airport"))
        )
        destination_button.click()
    except TimeoutException:
        raise AirportButtonError(f"Arrival option '{airport_code}' button not found or not clickable within the timeout.")
    time.sleep(1)

def find_month(driver, date):
    year = date[0]
    month = date[1]
    try:
        target_month_year_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//div[@class='month-select-wrap']//div[@data-date='" + month + "-" + year + "']"))
        )
        target_month_year_element.click()
    except TimeoutException:
        raise DateError(f"Target month '{month} {year}' not found or not clickable within the timeout.")
    time.sleep(1)


def find_day(driver, date):
    day = date[2]
    try:
        target_day_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//div[contains(@class, 'vc-day') and .//span[contains(@class, 'date') and normalize-space(text())='" + day + "']]"))
        )
        target_day_element.click()
    except TimeoutException:
        raise DateError(f"Target day '{day}' not found or not clickable within the timeout.")
    time.sleep(1)


def find_price(driver, departure_date, return_date):
    departure_date = departure_date.split("-")
    return_date = return_date.split("-")
    find_month(driver, departure_date)
    find_day(driver, departure_date)
    find_month(driver, return_date)
    find_day(driver, return_date)
    try:
        price_eur = driver.find_element(By.CLASS_NAME, "selected-price-main")
        price_cents = driver.find_element(By.CLASS_NAME, "selected-price-decimal")
        print (price_eur.text + "." + price_cents.text)
        try: 
            return (float(price_eur.text + "." + price_cents.text))
        except ValueError:
            return None
    except TimeoutException:
        # print(f"Target price is not found for these dates.")
        return None
    

def create_date_array():
    # Get today's date
    today = datetime.date.today()+datetime.timedelta(days=30)   
    days_tocheck = 3 
    date_array = [[None] + [today + datetime.timedelta(days=i) for i in range(days_tocheck)]]

    for i in range(1, days_tocheck*2+1):
        date_row = today + datetime.timedelta(days=i)
        row_data = [date_row] + [date_row > date_column and (date_row - date_column).days <= days_tocheck for date_column in date_array[0][1:]]
        date_array.append(row_data)

    return date_array


def process_date_cells(driver, date_array):
    for row_index, row in enumerate(date_array[1:]):
        date_row = row[0]
        for col_index, value in enumerate(row[1:]):
            if value:
                date_column = date_array[0][col_index + 1]
                date_array[row_index + 1][col_index + 1] = find_price(driver, date_column.strftime("%Y-%-m-%-d"), date_row.strftime("%Y-%-m-%-d"))
                



def print_date_array(date_array):
    for row in date_array:
        print("\t".join(map(lambda x: str(x) if x is not None else "", row)))  

def save_prices(array):
    wb = Workbook()
    ws = wb.active
    for row_data in array:
        ws.append(row_data)
    for col_num, cell in enumerate(ws[1], start=1):
        if col_num > 1:
            cell.value = str(cell.value)

    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, datetime.date):
                cell.value = str(cell.value)

    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None and cell.value is False:
                cell.value = 'x'

    wb.save('output.xlsx')


def compare_prices(start_day, departure_days_tocheck, arrivals_days_tocheck):
    # Creates an array of dates to check
    date_array = [[None] + [start_day + datetime.timedelta(days=i) for i in range(departure_days_tocheck)]]

    for i in range(1, arrivals_days_tocheck + departure_days_tocheck):
        date_row = start_day + datetime.timedelta(days=i)
        row_data = [date_row] + [date_row > date_column and (date_row - date_column).days <= arrivals_days_tocheck for date_column in date_array[0][1:]]
        date_array.append(row_data)
    # Calculates the price for each date
    for row_index, row in enumerate(date_array[1:]):
        date_row = row[0]
        for col_index, value in enumerate(row[1:]):
            if value:
                date_column = date_array[0][col_index + 1]
                date_array[row_index + 1][col_index + 1] = find_price(driver, date_column.strftime("%Y-%-m-%-d"), date_row.strftime("%Y-%-m-%-d"))
    # Saves prices to Excel
    wb = Workbook()
    ws = wb.active
    for row_data in date_array:
        ws.append(row_data)
    for col_num, cell in enumerate(ws[1], start=1):
        if col_num > 1:
            cell.value = str(cell.value)

    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, datetime.date):
                cell.value = str(cell.value)

    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None and cell.value is False:
                cell.value = 'x'

    wb.save('output.xlsx')

start_day = datetime.date.today() + datetime.timedelta(days=30)

try:
    accept_cookies(driver)
    set_departure_airport(driver, departure_airport)
    set_destination_airport(driver, destination_airport)
    compare_prices(start_day, 3, 3)

    # The following code will not be executed if an AirportButtonError occurs.
    # Place your additional logic here.

except (AirportButtonError, TimeoutException, Exception) as e:
    print(f"Error: {e}")
    # Add any cleanup code or exit the program as needed
finally:
    # Add any cleanup code here
    driver.quit()
