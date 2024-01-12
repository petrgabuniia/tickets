from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import time
import datetime
from openpyxl import Workbook
import tkinter as tk
from tkinter import ttk
from datetime import date
import os

class AirportButtonError(Exception):
    pass

class DateError(Exception):
    pass



def accept_cookies(driver): #Automaticall accepts cookies
    try:
        cookie_popup = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "cookie-alert"))
        )
        accept_button = cookie_popup.find_element(By.XPATH, "//button[contains(text(), 'Accept')]")
        accept_button.click()
    except TimeoutException:
        print("No cookie consent popup found or it was not displayed within the timeout.")

def set_departure_airport(driver, airport_code): #Automatically sets departure airport
    try:
        airport_input = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'fsf-typeahead') and contains(@class, 'is-index')][1]//input"))
        )
        airport_input.clear()
        airport_input.send_keys(airport_code)
        airport_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "airport"))
        )
        airport_button.click()
    except TimeoutException:
        raise AirportButtonError(f"Departure option '{airport_code}' button not found or not clickable within the timeout.")


def set_destination_airport(driver, airport_code): #Automatically sets destination airport
    destination_input = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'fsf-typeahead') and contains(@class, 'is-index')][2]//input"))
    )
    destination_input.clear()
    destination_input.send_keys(destination_airport)
    try:
        destination_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "airport"))
        )
        destination_button.click()
    except TimeoutException:
        raise AirportButtonError(f"Arrival option '{airport_code}' button not found or not clickable within the timeout.")
    time.sleep(1)

def find_month(driver, date): #Finds and selects the month of the flight 
    year = date[0]
    month = date[1]
    try:
        target_month_year_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//div[@class='month-select-wrap']//div[@data-date='" + month + "-" + year + "']"))
        )
        target_month_year_element.click()
    except TimeoutException:
        raise DateError(f"Target month '{month} {year}' not found or not clickable within the timeout.")
    time.sleep(1)


def find_day(driver, date): #Finds and selects the day of the flight 
    day = date[2]
    try:
        target_day_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//div[contains(@class, 'vc-day') and .//span[contains(@class, 'date') and normalize-space(text())='" + day + "']]"))
        )
        target_day_element.click()
    except TimeoutException:
        raise DateError(f"Target day '{day}' not found or not clickable within the timeout.")
    time.sleep(1)


def find_price(driver, departure_date, return_date): #Finds the price of the flight
    departure_date = departure_date.split("-")
    return_date = return_date.split("-")
    find_month(driver, departure_date)
    find_day(driver, departure_date)
    find_month(driver, return_date)
    find_day(driver, return_date)
    try:
        price_eur = driver.find_element(By.CLASS_NAME, "selected-price-main")
        price_cents = driver.find_element(By.CLASS_NAME, "selected-price-decimal")
        print (price_eur.text + "." + price_cents.text)
        try: 
            return (float(price_eur.text + "." + price_cents.text))
        except ValueError:
            return None
    except TimeoutException:

        return None
    

def create_date_array(): #Creates an array of dates to check
    today = datetime.date.today()+datetime.timedelta(days=30)   
    days_tocheck = 3 
    date_array = [[None] + [today + datetime.timedelta(days=i) for i in range(days_tocheck)]]

    for i in range(1, days_tocheck*2+1):
        date_row = today + datetime.timedelta(days=i)
        row_data = [date_row] + [date_row > date_column and (date_row - date_column).days <= days_tocheck for date_column in date_array[0][1:]]
        date_array.append(row_data)

    return date_array


def process_date_cells(driver, date_array): #Calculates the price for each date
    for row_index, row in enumerate(date_array[1:]):
        date_row = row[0]
        for col_index, value in enumerate(row[1:]):
            if value:
                date_column = date_array[0][col_index + 1]
                date_array[row_index + 1][col_index + 1] = find_price(driver, date_column.strftime("%Y-%-m-%-d"), date_row.strftime("%Y-%-m-%-d"))
                



def print_date_array(date_array): #Prints the array of dates
    for row in date_array:
        print("\t".join(map(lambda x: str(x) if x is not None else "", row)))  

def save_prices(array): #Saves the array with prices to Excel
    wb = Workbook()
    ws = wb.active
    for row_data in array:
        ws.append(row_data)
    for col_num, cell in enumerate(ws[1], start=1):
        if col_num > 1:
            cell.value = str(cell.value)

    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, datetime.date):
                cell.value = str(cell.value)

    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None and cell.value is False:
                cell.value = 'x'

    wb.save('output.xlsx')


def compare_prices(start_day, departure_days_tocheck, arrivals_days_tocheck): #Makes the whole process of comparing prices
    date_array = [[None] + [start_day + datetime.timedelta(days=i) for i in range(departure_days_tocheck)]]

    for i in range(1, arrivals_days_tocheck + departure_days_tocheck):
        date_row = start_day + datetime.timedelta(days=i)
        row_data = [date_row] + [date_row > date_column and (date_row - date_column).days <= arrivals_days_tocheck for date_column in date_array[0][1:]]
        date_array.append(row_data)
    # Calculates the price for each date
    for row_index, row in enumerate(date_array[1:]):
        date_row = row[0]
        for col_index, value in enumerate(row[1:]):
            if value:
                date_column = date_array[0][col_index + 1]
                date_array[row_index + 1][col_index + 1] = find_price(driver, date_column.strftime("%Y-%-m-%-d"), date_row.strftime("%Y-%-m-%-d"))
    wb = Workbook()
    ws = wb.active
    for row_data in date_array:
        ws.append(row_data)
    for col_num, cell in enumerate(ws[1], start=1):
        if col_num > 1:
            cell.value = str(cell.value)

    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, datetime.date):
                cell.value = str(cell.value)

    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None and cell.value is False:
                cell.value = 'x'

    wb.save('output.xlsx')

def get_date(): #Creates a window to enter the date
    def on_get_date():
        try:
            year, month, day = map(int, (year_entry.get(), month_entry.get(), day_entry.get()))
            selected_date = date(year, month, day)
            root.result = selected_date
            root.destroy()
        except ValueError:
            result_label.config(text="Incorrect date format. Please enter a valid date.")

    root = tk.Tk()
    root.title("Date to start checking prices from")

    year_entry = ttk.Entry(root)
    month_entry = ttk.Entry(root)
    day_entry = ttk.Entry(root)

    entries = {"Year:": year_entry, "Month:": month_entry, "Day:": day_entry}
    for i, (label_text, entry_widget) in enumerate(entries.items()):
        ttk.Label(root, text=label_text).grid(row=i, column=0, padx=5, pady=5, sticky="e")
        entry_widget.grid(row=i, column=1, padx=5, pady=5)

    ttk.Button(root, text="Get Date", command=on_get_date).grid(row=i + 1, column=0, columnspan=2, pady=10)

    result_label = ttk.Label(root, text="")
    result_label.grid(row=i + 2, column=0, columnspan=2, pady=5)

    root.mainloop()
    return getattr(root, 'result', None)

def get_string(message): #Creates a window to enter airport code
    def on_get_string():
        result_str = date_entry.get()
        root.result = result_str
        root.destroy()
    root = tk.Tk()
    root.title(message)
    date_entry = ttk.Entry(root)
    date_entry.grid(row=0, column=0, columnspan=2, padx=5, pady=5)
    ttk.Button(root, text="Done", command=on_get_string).grid(row=1, column=0, columnspan=2, pady=10)
    result_label = ttk.Label(root, text="")
    result_label.grid(row=2, column=0, columnspan=2, pady=5)
    root.mainloop()

    return getattr(root, 'result', None)
    


def get_integer(message): #Creates a window to enter the number of days
    def on_get_integer():
        try:
            result_int = int(entry.get())
            if 0 < result_int <= 25:
                root.result = result_int
                root.destroy()
            else:
                error_label.config(text="Please enter an integer between 1 and 25.")
        except ValueError:
            error_label.config(text="Please enter a valid integer.")

    root = tk.Tk()
    root.title(message)

    entry = ttk.Entry(root)
    entry.grid(row=0, column=0, columnspan=2, padx=5, pady=5)

    ttk.Button(root, text="Done", command=on_get_integer).grid(row=1, column=0, columnspan=2, pady=10)

    error_label = ttk.Label(root, text="")
    error_label.grid(row=2, column=0, columnspan=2, pady=5)

    root.mainloop()
    return getattr(root, 'result', None)


current_directory = os.path.dirname(os.path.abspath(__file__))
file_to_delete = "output.xlsx"
file_path = os.path.join(current_directory, file_to_delete)
if os.path.exists(file_path):
    os.remove(file_path)
    print(f"{file_to_delete} deleted successfully.")
else:
    print(f"{file_to_delete} not found in the directory.")



start_day = get_date()

# Your departure, destination and date details
departure_airport = get_string("Departure Airport Code")
departure_airport = departure_airport.upper()
departure_airport = departure_airport.replace(" ", "")
destination_airport = get_string("Destination Airport Code")
destination_airport = destination_airport.upper()
destination_airport = destination_airport.replace(" ", "")
depatrure_days_tocheck = get_integer("How many days to check for departure?")
arrival_days_tocheck = get_integer("What is the maksimal length of your trip?")


options = webdriver.ChromeOptions()
driver = webdriver.Chrome(options=options)
url = "https://www.airbaltic.com/en-LV/index"
driver.get(url)

try:
    accept_cookies(driver)
    set_departure_airport(driver, departure_airport)
    set_destination_airport(driver, destination_airport)
    compare_prices(start_day, depatrure_days_tocheck, arrival_days_tocheck)

except (AirportButtonError, TimeoutException, Exception) as e:
    print(f"Error: {e}")

finally:
    driver.quit()

